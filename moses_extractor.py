#!/usr/bin/env python3
"""
MOSES Value Extraction Tool - Enhanced Version
Extracts damage stability data matching Table 10.4.2 format
"""

import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def extract_moses_data(file_path):
    """
    Extract complete damage stability data from MOSES output file
    
    Returns:
        list: List of dictionaries containing all required stability data
    """
    with open(file_path, 'r') as f:
        content = f.read()
    
    results = []
    lines = content.split('\n')
    
    # Process in two passes: first collect stability data, then match with drafts
    stability_data = {}
    draft_data = {}
    
    current_case = None
    in_stability_summary = False
    in_draft_marks = False
    temp_heel = None
    temp_trim = None
    temp_drafts = {}
    
    for i, line in enumerate(lines):
        line_stripped = line.strip()
        
        # Detect damage case from section headers or VCG line
        if 'DAMAGE STABILITY Case-' in line_stripped:
            match = re.search(r'Case-(\d+)\s*\(Compartment\s+(\w+)\s+Flooded\)', line_stripped)
            if match:
                current_case = match.group(1)
        elif 'INTACT TOW CONDITION' in line_stripped:
            current_case = 'Intact'
        elif 'Damage = NONE' in line_stripped and 'VCG' in line_stripped:
            current_case = 'Intact'
        elif 'Damage =' in line_stripped and 'VCG' in line_stripped:
            # Extract case from "Damage = XPO" format
            match = re.search(r'Damage = (\w+)\s', line_stripped)
            if match:
                damage_id = match.group(1)
                if damage_id != 'NONE!':
                    # Extract case number from damage ID (e.g., "1PO" -> "1")
                    case_match = re.search(r'(\d+)', damage_id)
                    if case_match:
                        current_case = case_match.group(1)
        
        # Detect draft mark readings section
        if '+++ D R A F T   M A R K   R E A D I N G S +++' in line_stripped:
            in_draft_marks = True
            temp_drafts = {}
            continue
        
        # Parse draft marks
        if in_draft_marks and 'AFT(P)' in line_stripped and 'AFT(S)' in line_stripped:
            draft_pattern = r'(\w+\([PS]\))\s+([\d.]+)'
            matches = re.findall(draft_pattern, line_stripped)
            
            for name, value in matches:
                if name not in ['MEAN(P)', 'MEAN(S)']:
                    temp_drafts[name] = float(value)
            
            if current_case and temp_drafts:
                draft_data[current_case] = temp_drafts.copy()
            
            in_draft_marks = False
        
        # Detect stability summary section
        if '+++ S T A B I L I T Y   S U M M A R Y +++' in line_stripped:
            in_stability_summary = True
            temp_heel = None
            temp_trim = None
            continue
        
        # Extract Roll (Heel) and Pitch (Trim)
        if in_stability_summary:
            if 'Roll' in line_stripped and '=' in line_stripped and 'Deg' in line_stripped:
                roll_match = re.search(r'Roll\s*=\s*([-\d.]+)\s*Deg', line_stripped)
                if roll_match:
                    temp_heel = float(roll_match.group(1))
            
            if 'Pitch' in line_stripped and '=' in line_stripped and 'Deg' in line_stripped:
                pitch_match = re.search(r'Pitch\s*=\s*([-\d.]+)\s*Deg', line_stripped)
                if pitch_match:
                    temp_trim = float(pitch_match.group(1))
        
        # Extract Area Ratio
        if in_stability_summary and 'Area Ratio' in line_stripped and 'Passes' in line_stripped:
            match = re.search(r'Area Ratio\s+>=\s+([\d.]+)\s+([\d.]+)\s+Passes', line_stripped)
            if match:
                required_ratio = float(match.group(1))
                actual_ratio = float(match.group(2))
                
                if current_case:
                    stability_data[current_case] = {
                        'heel': temp_heel if temp_heel is not None else 0.0,
                        'trim': temp_trim if temp_trim is not None else 0.0,
                        'area_ratio_actual': actual_ratio,
                        'area_ratio_required': required_ratio
                    }
                
                in_stability_summary = False
    
    # Now combine stability data with draft data
    for case in sorted(stability_data.keys(), key=lambda x: (x != 'Intact', int(x) if x.isdigit() else 0)):
        if case in draft_data:
            results.append({
                'case': case,
                'drafts': draft_data[case],
                'heel': stability_data[case]['heel'],
                'trim': stability_data[case]['trim'],
                'area_ratio_actual': stability_data[case]['area_ratio_actual'],
                'area_ratio_required': stability_data[case]['area_ratio_required'],
                'remarks': 'Pass'
            })
    
    return results

def create_excel_report(data, output_path):
    """
    Create formatted Excel report matching Table 10.4.2 format
    
    Args:
        data: List of dictionaries with stability information
        output_path: Path to save the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Damage Stability Results"
    
    # Define styles
    title_font = Font(name='Arial', size=14, bold=True)
    title_alignment = Alignment(horizontal='center', vertical='center')
    
    header_font = Font(name='Arial', size=10, bold=True, color='000000')
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    data_font = Font(name='Arial', size=10)
    data_alignment = Alignment(horizontal='center', vertical='center')
    
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Title
    ws.merge_cells('A1:J1')
    ws['A1'] = 'Table 10.4.2 Damage Stability Results'
    ws['A1'].font = title_font
    ws['A1'].alignment = title_alignment
    ws.row_dimensions[1].height = 25
    
    # Main headers (Row 2)
    main_headers = [
        ('A2', 'Case\nNo.', 1),
        ('B2', 'Draft Mark (m)', 4),
        ('F2', 'Heel\n(deg)', 1),
        ('G2', 'Trim\n(deg)', 1),
        ('H2', 'Wind Area Ratio', 2),
        ('J2', 'Remarks', 1)
    ]
    
    for cell_ref, text, merge_cols in main_headers:
        if merge_cols > 1:
            start_col = cell_ref[0]
            end_col = chr(ord(start_col) + merge_cols - 1)
            ws.merge_cells(f'{start_col}2:{end_col}2')
        
        cell = ws[cell_ref]
        cell.value = text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Sub headers (Row 3)
    sub_headers = [
        ('A3', ''),
        ('B3', 'Aft\nPort'),
        ('C3', 'Aft\nStbd'),
        ('D3', 'Fwd\nPort'),
        ('E3', 'Fwd\nStbd'),
        ('F3', ''),
        ('G3', ''),
        ('H3', 'Actual'),
        ('I3', 'Required'),
        ('J3', '')
    ]
    
    for cell_ref, text in sub_headers:
        cell = ws[cell_ref]
        cell.value = text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 25
    
    # Data rows
    for row_num, record in enumerate(data, 4):
        # Case number
        cell = ws.cell(row=row_num, column=1)
        cell.value = record['case']
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = border
        
        # Draft marks
        draft_mapping = {
            2: 'AFT(P)',
            3: 'AFT(S)',
            4: 'FWD(P)',
            5: 'FWD(S)'
        }
        
        for col_num, draft_name in draft_mapping.items():
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = record['drafts'].get(draft_name, None)
            if cell.value is not None:
                cell.number_format = '0.00'
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border
        
        # Heel (deg)
        cell = ws.cell(row=row_num, column=6)
        cell.value = record['heel']
        cell.number_format = '0.00'
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = border
        
        # Trim (deg)
        cell = ws.cell(row=row_num, column=7)
        cell.value = record['trim']
        cell.number_format = '0.00'
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = border
        
        # Wind Area Ratio - Actual
        cell = ws.cell(row=row_num, column=8)
        cell.value = record['area_ratio_actual']
        cell.number_format = '0.00'
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = border
        
        # Wind Area Ratio - Required
        cell = ws.cell(row=row_num, column=9)
        cell.value = record['area_ratio_required']
        cell.number_format = '0.0'
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = border
        
        # Remarks
        cell = ws.cell(row=row_num, column=10)
        cell.value = record['remarks']
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = border
    
    # Column widths
    column_widths = [8, 10, 10, 10, 10, 8, 8, 10, 10, 10]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Save workbook
    wb.save(output_path)
    print(f"Excel file created successfully: {output_path}")

def main():
    """Main execution function"""
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python moses_extractor_v2.py <input_file> [output_file]")
        print("Example: python moses_extractor_v2.py out00001.txt damage_stability_results.xlsx")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else 'Damage_Stability_Results.xlsx'
    
    print(f"Processing MOSES file: {input_file}")
    
    # Extract data
    data = extract_moses_data(input_file)
    
    print(f"Extracted {len(data)} damage cases")
    
    # Display extracted data
    print("\nExtracted Data:")
    print("-" * 120)
    print(f"{'Case':<8} {'Aft Port':<10} {'Aft Stbd':<10} {'Fwd Port':<10} {'Fwd Stbd':<10} {'Heel':<8} {'Trim':<8} {'Area Ratio':<12} {'Remarks'}")
    print("-" * 120)
    for record in data:
        aft_p = f"{record['drafts'].get('AFT(P)', 0):.2f}" if 'AFT(P)' in record['drafts'] else 'N/A'
        aft_s = f"{record['drafts'].get('AFT(S)', 0):.2f}" if 'AFT(S)' in record['drafts'] else 'N/A'
        fwd_p = f"{record['drafts'].get('FWD(P)', 0):.2f}" if 'FWD(P)' in record['drafts'] else 'N/A'
        fwd_s = f"{record['drafts'].get('FWD(S)', 0):.2f}" if 'FWD(S)' in record['drafts'] else 'N/A'
        
        print(f"{record['case']:<8} {aft_p:<10} {aft_s:<10} {fwd_p:<10} {fwd_s:<10} "
              f"{record['heel']:>7.2f} {record['trim']:>7.2f} {record['area_ratio_actual']:>11.2f} {record['remarks']}")
    print("-" * 120)
    
    # Create Excel report
    create_excel_report(data, output_file)

if __name__ == '__main__':
    main()
